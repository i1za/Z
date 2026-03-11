import argparse
import json
import os
import sys
from datetime import date, datetime
from pathlib import Path
from urllib import error, request

from openpyxl import load_workbook

DEFAULT_SUPABASE_URL = "https://rgbzximweiafgdukppbf.supabase.co"
DEFAULT_SUPABASE_ANON_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InJnYnp4aW13"
    "ZWlhZmdkdWtwcGJmIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzMwNDU5MDcsImV4cCI6MjA4ODYyMTkw"
    "N30.fMGcuP-E2mxG_LHCq4TLaI8087pi17oIMoQuNlNspUs"
)

EMPLOYEE_COLUMNS = [
    "sn",
    "full_name",
    "company_name",
    "job_title",
    "nationality",
    "phone",
    "branch_name",
    "residence_status",
    "start_date",
    "basic_salary",
    "allowances",
    "incentives",
    "total_salary",
    "payment_method",
    "insurance_company",
    "insurance_coverage",
    "food_permit_emirate",
    "food_permit_expiry",
    "health_card_emirate",
    "health_card_expiry",
    "passport_expiry",
    "passport_held_by",
    "residence_expiry",
    "work_card_expiry",
]


def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def to_date_iso(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None

    formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def to_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def to_int(value):
    num = to_number(value)
    if num is None:
        return None
    try:
        return int(num)
    except ValueError:
        return None


def parse_main_sheet(ws):
    records = []
    for row in range(5, ws.max_row + 1):
        full_name = clean_text(ws.cell(row=row, column=2).value)
        if not full_name:
            continue

        passport_holder = (
            clean_text(ws.cell(row=row, column=29).value)
            or clean_text(ws.cell(row=row, column=28).value)
        )

        record = {
            "sn": to_int(ws.cell(row=row, column=1).value),
            "full_name": full_name,
            "company_name": clean_text(ws.cell(row=row, column=3).value),
            "job_title": clean_text(ws.cell(row=row, column=4).value),
            "nationality": clean_text(ws.cell(row=row, column=5).value),
            "phone": clean_text(ws.cell(row=row, column=6).value),
            "branch_name": clean_text(ws.cell(row=row, column=7).value),
            "residence_status": clean_text(ws.cell(row=row, column=8).value),
            "start_date": to_date_iso(ws.cell(row=row, column=9).value),
            "work_card_expiry": to_date_iso(ws.cell(row=row, column=10).value),
            "residence_expiry": to_date_iso(ws.cell(row=row, column=11).value),
            "basic_salary": to_number(ws.cell(row=row, column=12).value),
            "allowances": to_number(ws.cell(row=row, column=13).value),
            "incentives": to_number(ws.cell(row=row, column=14).value),
            "total_salary": to_number(ws.cell(row=row, column=15).value),
            "payment_method": clean_text(ws.cell(row=row, column=16).value),
            "insurance_company": clean_text(ws.cell(row=row, column=17).value),
            "insurance_coverage": clean_text(ws.cell(row=row, column=18).value),
            "food_permit_emirate": clean_text(ws.cell(row=row, column=20).value),
            "food_permit_expiry": to_date_iso(ws.cell(row=row, column=21).value),
            "health_card_emirate": clean_text(ws.cell(row=row, column=23).value),
            "health_card_expiry": to_date_iso(ws.cell(row=row, column=24).value),
            "passport_expiry": to_date_iso(
                ws.cell(row=row, column=26).value
                or ws.cell(row=row, column=27).value
            ),
            "passport_held_by": passport_holder,
        }
        records.append(record)
    return records


def parse_cancellation_sheet(ws):
    records = []
    for row in range(4, ws.max_row + 1):
        full_name = clean_text(ws.cell(row=row, column=2).value)
        if not full_name:
            continue

        record = {
            "sn": to_int(ws.cell(row=row, column=1).value),
            "full_name": full_name,
            "company_name": "سجل الكنسلة",
            "job_title": clean_text(ws.cell(row=row, column=3).value),
            "nationality": clean_text(ws.cell(row=row, column=4).value),
            "phone": clean_text(ws.cell(row=row, column=5).value),
            "start_date": to_date_iso(ws.cell(row=row, column=6).value),
            "residence_expiry": to_date_iso(ws.cell(row=row, column=7).value),
            "residence_status": clean_text(ws.cell(row=row, column=8).value) or "ملغي",
        }
        records.append(record)
    return records


def dedupe_records(records):
    seen = set()
    unique = []
    for record in records:
        key = (
            record.get("full_name"),
            record.get("job_title"),
            record.get("start_date"),
            record.get("company_name"),
            record.get("residence_status"),
        )
        if key in seen:
            continue
        seen.add(key)
        unique.append(record)
    return unique


def http_request(method, url, anon_key, payload=None):
    headers = {
        "apikey": anon_key,
        "Authorization": f"Bearer {anon_key}",
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    data = None if payload is None else json.dumps(payload, ensure_ascii=False).encode("utf-8")

    req = request.Request(url, data=data, headers=headers, method=method)
    try:
        with request.urlopen(req, timeout=60) as resp:
            return resp.status, resp.read().decode("utf-8")
    except error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"{method} {url} failed ({exc.code}): {detail}") from exc


def load_source_records(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    if not wb.worksheets:
        raise RuntimeError("Workbook has no sheets")

    main_sheet = wb.worksheets[0]
    records = parse_main_sheet(main_sheet)

    if len(wb.worksheets) > 1:
        records.extend(parse_cancellation_sheet(wb.worksheets[1]))

    return dedupe_records(records)


def run_import(excel_path, supabase_url, anon_key, replace_existing):
    source_records = load_source_records(excel_path)
    records = [{column: item.get(column) for column in EMPLOYEE_COLUMNS} for item in source_records]
    if not records:
        raise RuntimeError("No employee records found in workbook")

    base = supabase_url.rstrip("/")
    if replace_existing:
        delete_url = f"{base}/rest/v1/employees?id=not.is.null"
        http_request("DELETE", delete_url, anon_key)

    batch_size = 200
    for i in range(0, len(records), batch_size):
        batch = records[i : i + batch_size]
        insert_url = f"{base}/rest/v1/employees"
        http_request("POST", insert_url, anon_key, batch)

    return len(records)


def resolve_excel_path(input_path):
    if input_path:
        p = Path(input_path)
        if p.exists():
            return p
        raise RuntimeError(f"Excel file not found: {input_path}")

    default = Path(r"C:\Users\Acer\Documents\employees_source.xlsx")
    if default.exists():
        return default

    docs = Path.home() / "Documents"
    if docs.exists():
        candidates = sorted(
            docs.glob("*.xlsx"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        preferred = [
            p
            for p in candidates
            if "موظفين" in p.name or "employee" in p.name.lower() or "employees" in p.name.lower()
        ]
        if preferred:
            return preferred[0]
        if candidates:
            return candidates[0]

    raise RuntimeError("Provide --excel path to the source workbook")


def main():
    parser = argparse.ArgumentParser(description="Import Al Eatemad employee Excel data into Supabase")
    parser.add_argument("--excel", help="Path to source workbook (.xlsx)")
    parser.add_argument("--replace", action="store_true", help="Replace all existing rows before insert")
    args = parser.parse_args()

    excel_path = resolve_excel_path(args.excel)
    supabase_url = os.getenv("VITE_SUPABASE_URL", DEFAULT_SUPABASE_URL).strip()
    anon_key = os.getenv("VITE_SUPABASE_ANON_KEY", DEFAULT_SUPABASE_ANON_KEY).strip()
    service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip() or os.getenv(
        "VITE_SUPABASE_SERVICE_ROLE_KEY", ""
    ).strip()
    key = service_key or anon_key

    if not supabase_url or not key:
        raise RuntimeError("Missing Supabase configuration")

    imported_count = run_import(
        excel_path=excel_path,
        supabase_url=supabase_url,
        anon_key=key,
        replace_existing=args.replace,
    )
    print(f"Imported {imported_count} employee records from: {excel_path}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}")
        sys.exit(1)
